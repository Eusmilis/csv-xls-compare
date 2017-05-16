from common import LineDifferencesInfo, DifferencesInfo, generate_report
from openpyxl import load_workbook
from pprint import pprint

def letter_to_index(letter):
    letter = letter.upper()
    result = 0
    for index, char in enumerate(reversed(letter)):
        num = ord(char)-64
        final_num = (26**index)*num
        result += final_num
    return result

def get_data(filename, *key_columns):
    workbook = load_workbook(filename)
    worksheet = workbook.worksheets[0]
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    data = {}
    for row in range(1, max_row+1):
        values = [worksheet.cell(row=row, column=column).value for column in range(1, max_column+1)]
        data[tuple(worksheet.cell(row=row, column=letter_to_index(col)).value for col in key_columns)] = values
    return data
